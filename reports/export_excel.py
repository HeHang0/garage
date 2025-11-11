from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

columns_map = {
    'Year': '年',
    'YearMonth': '年月',
    'Month': '月',
    'Fee': '金额',
    'CPH': '车牌号',
    'TypeClass': '类型',
    'VisitCount': '进出次数',
    'CPHCount': '车辆数',
    'DayInCount': '白天进入次数',
    'DayCount': '白天出入次数',
    'DayOutCount': '白天离开次数',
    'NightInCount': '夜间进入次数',
    'NightCount': '夜间出入次数',
    'NightOutCount': '夜间离开次数',
    'MaxStayHour': '最长停放小时',
    'MaxStayDay': '最长停放天数',
    'UserName': '名称',
    'HomeAddress': '地址',
    'Province': '地区',
    'Count': '数量',
    'CarType': '车辆类型',
    'StayText': '停车时长',
    'InTime': '进入时间',
    'OutTime': '离开时间',
    'InCount': '进入次数',
    'OutCount': '离开次数',
    'InGateName': '进入位置',
    'OutGateName': '离开位置',
    'IsTenant': '是否租户',
    'CouponCPH': '问券车牌号',
    'HasCoupon': '是否参加问券',
    'CouponUserName': '问券姓名',
    'UserNameConsistency': '姓名一致性'
}

def export_to_excel(df, writer, sheet_name="Sheet1"):
    df.rename(columns=columns_map).to_excel(writer, index=False, sheet_name=sheet_name)

def excel_auto_width(excel_path, simple=True):
    wb = load_workbook(excel_path)

    for sheet in wb.worksheets:
        if simple:
            first_row = next(sheet.iter_rows(min_row=1, max_row=1))  # 第一行
            for cell in first_row:
                if cell.value:
                    val = str(cell.value)
                    # 中文宽度加权处理：中文 2，英文/数字 1
                    length = sum(2 if ord(c) > 127 else 1 for c in val)
                    col_letter = get_column_letter(cell.column)
                    sheet.column_dimensions[col_letter].width = length + 2  # 加点缓冲
            continue
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    if cell.value:
                        val = str(cell.value)
                        # 中文字符更宽，这里按比例增加
                        length = sum(2 if ord(c) > 127 else 1 for c in val)
                        max_length = max(max_length, length)
                except:
                    pass
            adjusted_width = max_length + 2  # 加一点缓冲宽度
            sheet.column_dimensions[column_letter].width = adjusted_width

    # 保存修改后的 Excel
    wb.save(excel_path)
